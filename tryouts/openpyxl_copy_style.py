from openpyxl.styles import numbers, Font, PatternFill

import openpyxl
from copy import copy

file_path = 'test_copy_color.xlsx'
new = "new"

wb = openpyxl.load_workbook(file_path)
wb.remove(wb[new]) if new in wb.sheetnames else None
ns = wb.create_sheet(new)

src_sheet = wb['L1']
cell_fill = PatternFill()
for row in src_sheet.rows:
    for cell in row:
        new_cell = ns.cell(row=cell.row, column=cell.col_idx, value=cell.value)
        cell_fill = copy(cell.fill)
        if cell.has_style:
            # new_cell.font = copy(cell.font)
            # new_cell.border = copy(cell.border)
            new_cell.fill = cell_fill
            # new_cell.fill = copy(cell.fill)
            # new_cell.number_format = copy(cell.number_format)
            # new_cell.protection = copy(cell.protection)
            # new_cell.alignment = copy(cell.alignment)
print(cell_fill)

print(cell_fill.__class__.__name__)

wb.save(file_path)
