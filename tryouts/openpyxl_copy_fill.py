from openpyxl.worksheet import worksheet
from openpyxl.utils.cell import column_index_from_string
from openpyxl.styles import numbers, Font

from openpyxl import load_workbook
import openpyxl
from copy import copy

treeData = [["Type", "Leaf Color", "Height"], ["Maple", "Red", 549], ["Oak", "Green", 783], ["Pine", "Green", 1204]]
file_path = 'test_copy_color.xlsx'

wb = load_workbook(file_path)
# ws = wb["L_new"] # wb.create_sheet("L_new")
#
# for row in treeData:
#     ws.append(row)
Colors = openpyxl.styles.colors.COLOR_INDEX
print(Colors)


ds = wb['L1']
for row in ds.rows:
    for cell in row:

        print(f"{cell.value = } {cell.fill.fgColor.rgb = } ")
    print()

        # print(f"{cell.value = } {cell.font.color.index = } {cell.fill.start_color.index = } {cell.fill.__dict__.keys() = }")
        # result = Colors[ic]
        # print(f"{cell.value = } {result = }")


        # print(f"{cell.value = } {cell.font.color.__repr__ = } {cell.font.color.__dict__:}") # {cell.font.color.index = } {cell.font.color = }
        # print(f"{cell.value = } {cell.font.color.value = }")





        # f = cell.font
        # c = f.color
        # v = cell.value
        # print(f"{c=} {v=} {c.rgb=}")
#
# if cl.font.color != None and type(cl.font.color.rgb) == str:
#    # where cl =  cell of interest
#    rbg = cl.font.color.rgb


# wb.save(file_path)






#
#
# ws = wb['L1']
#
#
# ns = wb.create_sheet("L_new")
# ds = wb['L1']
#
#
# for row in ds.rows:
#     for cell in row:
#         new_cell = ns.cell(row=cell.row, column=cell.col_idx, value=cell.value)
#         if cell.has_style:
#             new_cell.font = copy(cell.font)
#             new_cell.border = copy(cell.border)
#             new_cell.fill = copy(cell.fill)
#             new_cell.number_format = copy(cell.number_format)
#             new_cell.protection = copy(cell.protection)
#             new_cell.alignment = copy(cell.alignment)
#
# wb.save('test2.xlsx')
