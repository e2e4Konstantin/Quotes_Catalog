from openpyxl.worksheet import worksheet
from openpyxl.utils.cell import column_index_from_string

from settings import Catalog, items_fonts, item_index, headers
from catalog import get_numeric_stamp
from excelutils.quotes_output import quotes_output
from openpyxl.styles import numbers


def _table_line_output(code: str, catalog: Catalog, sheet: worksheet, row: int) -> int:
    """ Записывает информацию из каталога о Таблице с шифром code на лист sheet в строку row. """
    table = catalog.tables[code]
    # две строки на таблицу
    row += 1
    sheet.cell(row=row, column=column_index_from_string('A')).value = table.chapter
    sheet.cell(row=row, column=column_index_from_string('B')).value = table.collection
    sheet.cell(row=row, column=column_index_from_string('C')).value = table.section
    sheet.cell(row=row, column=column_index_from_string('D')).value = table.subsection
    sheet.cell(row=row, column=column_index_from_string('E')).value = table.code
    sheet.cell(row=row, column=column_index_from_string('F')).value = ""    #table.code
    sheet.cell(row=row, column=column_index_from_string('G')).value = table.title
    # группировка 5
    group_number = item_index['table'] + 1
    sheet.row_dimensions.group(row, row + 1, outline_level=group_number)
    columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G']
    for column in columns:
        sheet.cell(row=row, column=column_index_from_string(column)).style = 'table_name'
        sheet.cell(row=row, column=column_index_from_string(column)).font = items_fonts[item_index['table']]
        sheet.cell(row=row, column=column_index_from_string(column)).number_format = numbers.FORMAT_TEXT

    column = column_index_from_string('K')
    sheet.cell(row=row - 1, column=column).value = headers['K1']
    sheet.cell(row=row - 1, column=column).style = 'further_quotes'
    sheet.merge_cells(start_row=row - 1, start_column=column, end_row=row - 1, end_column=column + 2)

    sheet.cell(row=row, column=column_index_from_string('K')).value = headers['K']
    sheet.cell(row=row, column=column_index_from_string('L')).value = headers['L']
    sheet.cell(row=row, column=column_index_from_string('M')).value = headers['M']
    for column in ['K', 'L', 'M']:
        sheet.cell(row=row, column=column_index_from_string(column)).style = 'further_quotes'

    column = column_index_from_string('N')
    sheet.cell(row=row - 1, column=column).value = headers['N1']
    sheet.cell(row=row - 1, column=column).style = 'title_attributes'

    attributes = []
    if table.attributes is not None:
        attributes.extend(table.attributes.split(','))
        attributes_length = len(attributes)
        for i, attribute in enumerate(attributes):
            sheet.cell(row=row, column=column + i).value = attribute
            sheet.cell(row=row, column=column + i).style = 'title_attributes'
        if attributes_length > 1:
            sheet.merge_cells(start_row=row - 1, start_column=column, end_row=row - 1,
                              end_column=column + attributes_length - 1)
    else:
        sheet.cell(row=row, column=column_index_from_string('N')).value = headers['N']
        sheet.cell(row=row, column=column_index_from_string('O')).value = headers['O']
        sheet.cell(row=row, column=column_index_from_string('N')).style = 'title_attributes'
        sheet.cell(row=row, column=column_index_from_string('O')).style = 'title_attributes'
        sheet.merge_cells(start_row=row - 1, start_column=column, end_row=row - 1, end_column=column + 1)

    parameters = []
    if table.parameters:
        column += len(attributes)
        parameters.extend(table.parameters.split(','))  # параметры
        parameters_length = len(parameters)
        if parameters_length > 0:
            mini_table_header = headers['P:T']              # таблица: 'от', 'до', 'ед.изм.', 'шаг', 'тип'
            parameter_tile_length = len(mini_table_header)

            for parameter in parameters:
                sheet.cell(row=row - 1, column=column).value = parameter  # название параметра
                sheet.cell(row=row - 1, column=column).style = 'title_parameter'

                i = 0
                for item in mini_table_header:
                    sheet.cell(row=row, column=column + i).value = item
                    sheet.cell(row=row, column=column + i).style = 'title_parameter'
                    i += 1

                column_delta = parameter_tile_length - 1
                sheet.merge_cells(start_row=row - 1, start_column=column, end_row=row - 1,
                                  end_column=column + column_delta)
                column += column_delta + 1

    return row + 1


def _get_tables(catalog: Catalog = None,
                chapter: str = None, collection: str = None,
                section: str = None, subsection: str = None) -> list[str] | None:
    table_codes = [code for code in catalog.tables.keys()
                   if catalog.tables[code].chapter == chapter and
                   catalog.tables[code].collection == collection and
                   catalog.tables[code].section == section and
                   catalog.tables[code].subsection == subsection
                   ]
    if len(table_codes) > 0:
        return sorted(table_codes, key=lambda x: tuple(map(int, get_numeric_stamp(x, 'table'))))
    return None


def tables_output(sheet: worksheet, catalog: Catalog, start_line: int,
                  chapter: str = None, collection: str = None,
                  section: str = None, subsection: str = None) -> int:
    """ Записывает информацию из каталога о таблицах на лист sheet начиная со строки start_line. """
    tables = _get_tables(catalog, chapter, collection, section, subsection)
    if tables:
        row = start_line
        for table in tables:
            # print('\t\t\t\t', f"{catalog.tables[table].code!r} {catalog.tables[table].title}")
            row = _table_line_output(table, catalog, sheet, row)
            row = quotes_output(sheet, catalog, row, chapter, collection, section, subsection, table)
        return row
    # print(f"\t\t\t\tнет ни одной таблицы")
    return start_line
