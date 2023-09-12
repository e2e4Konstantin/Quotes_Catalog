from openpyxl.worksheet import worksheet
from openpyxl.styles import Font, Alignment
from openpyxl.utils.cell import column_index_from_string
from openpyxl.styles import numbers

from settings import Catalog, items_fonts, item_index
from catalog import get_numeric_stamp
from read_quotes_parametes import (
    get_all_attributes_for_quote_from_data_by_index, get_all_parameters_for_quote_from_data_by_index
)


def _get_parameter_from_tuple(src_data: tuple[str, str, str, str, str]) -> dict:
    """ Преобразует кортеж со значениями параметра в словарь """
    # src_data = ("left", "right", "measure", "step", "type")
    p_value = dict()
    p_value["left"], p_value["right"], p_value["measure"], p_value["step"], p_value["type"] = src_data
    return p_value


def put_parameters_to_sheet(sheet: worksheet, parameters: list, row: int, table_parameters: str,
                            attribute_counter: int):
    """ На лист sheet в строку row, выводит значение атрибутов.  """
    # column_names = ("idx", "quote", "name", "left", "right", "measure", "step", "type")
    if len(parameters) > 0:
        table_parameters_list = [x.strip() for x in table_parameters.split(',')]  # список параметров из шапки таблицы
        # делаем словарь параметр: значения
        parameters_dict = {x[2]: _get_parameter_from_tuple(x[3:]) for x in parameters}
        # колонка равна стартовой для атрибутов + кол-во атрибутов
        start_column_number = column_index_from_string('N') + attribute_counter
        step = 0
        # перебираем параметры в шапке таблицы
        for i, headers_parameter in enumerate(table_parameters_list):
            parameter_value = parameters_dict.get(headers_parameter, None)
            if parameter_value:
                column = i + step
                sheet.cell(row=row, column=start_column_number + column).value = parameter_value.get("left", "0")
                sheet.cell(row=row, column=start_column_number + column + 1).value = parameter_value.get("right", "0")
                sheet.cell(row=row, column=start_column_number + column + 2).value = parameter_value.get("measure", "")
                sheet.cell(row=row, column=start_column_number + column + 3).value = parameter_value.get("step", "0")
                sheet.cell(row=row, column=start_column_number + column + 4).value = parameter_value.get("type", "0")
                step += 4
                for c in range(i + step + 1):
                    sheet.cell(row=row, column=start_column_number + c).style = 'line_table'



def _quote_line_output(code: str, catalog: Catalog, sheet: worksheet, row: int) -> int:
    """ Записывает информацию из каталога о Расценке с шифром code на лист sheet в строку row. """
    quote = catalog.quotes[code]
    # группировка 6
    group_number = item_index['quote'] + 1
    sheet.row_dimensions.group(row, row + 1, outline_level=group_number)

    sheet.cell(row=row, column=column_index_from_string('A')).value = quote.chapter
    sheet.cell(row=row, column=column_index_from_string('B')).value = quote.collection
    sheet.cell(row=row, column=column_index_from_string('C')).value = quote.section
    sheet.cell(row=row, column=column_index_from_string('D')).value = quote.subsection
    sheet.cell(row=row, column=column_index_from_string('E')).value = quote.table
    sheet.cell(row=row, column=column_index_from_string('F')).value = quote.code        # код расценки
    sheet.cell(row=row, column=column_index_from_string('G')).value = quote.title       # текст расценки
    sheet.cell(row=row, column=column_index_from_string('H')).value = quote.measure     # измеритель
    sheet.cell(row=row, column=column_index_from_string('I')).value = quote.stat        # статистика
    sheet.cell(row=row, column=column_index_from_string('J')).value = quote.flag
    sheet.cell(row=row, column=column_index_from_string('K')).value = quote.basic_slave     # тип расценки
    sheet.cell(row=row, column=column_index_from_string('L')).value = quote.link_cod        # код родительской расценки

    for column in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
        sheet.cell(row=row, column=column_index_from_string(column)).style = 'quote_line'
        sheet.cell(row=row, column=column_index_from_string(column)).font = items_fonts[item_index['quote']]
        sheet.cell(row=row, column=column_index_from_string(column)).number_format = numbers.FORMAT_TEXT

    sheet.cell(row=row, column=column_index_from_string('I')).alignment = Alignment(horizontal='right')
    sheet.cell(row=row, column=column_index_from_string('J')).alignment = Alignment(horizontal='center')

    quote_code_font = Font(name='Calibri', bold=True, size=8, color="000000")
    measure_font = Font(name='Calibri', bold=False, size=8, color="60497A")
    sheet.cell(row=row, column=column_index_from_string('F')).font = quote_code_font
    sheet.cell(row=row, column=column_index_from_string('H')).font = measure_font

    attributes = get_all_attributes_for_quote_from_data_by_index(code)  # получаем атрибуты для расценки из датасета
    table_attributes = catalog.tables[quote.table].attributes

    if len(attributes) > 0:
        table_attributes_list = [x.strip() for x in table_attributes.split(',')]  # список атрибутов из шапки таблицы
        attribute_titles = {x[2].strip(): x[3].strip() for x in attributes}  # словарь атрибут:значение
        start_column_number = column_index_from_string('N')
        for i, headers_attributes in enumerate(table_attributes_list):
            value_out = attribute_titles.get(headers_attributes, " ")
            sheet.cell(row=row, column=start_column_number + i).value = value_out
            sheet.cell(row=row, column=start_column_number + i).style = 'line_table'

    parameters = get_all_parameters_for_quote_from_data_by_index(code)  # получаем параметры для расценки из датасета
    table_parameters = catalog.tables[quote.table].parameters

    attribute_counter = len([x.strip() for x in table_attributes.split(',')])

    put_parameters_to_sheet(sheet, parameters, row, table_parameters, attribute_counter)






    return row+1


def _get_quotes(catalog: Catalog = None,
                chapter: str = None, collection: str = None,
                section: str = None, subsection: str = None, table: str = None) -> list[str] | None:
    """ Создает список шифров расценок и сортирует его. """
    tmp = catalog.quotes.keys()
    quote_codes = []
    for code in tmp:
        quote = catalog.quotes[code]
        if quote.chapter == chapter:
            if quote.collection == collection:
                if quote.section == section:
                    if quote.subsection == subsection:
                        if quote.table == table:
                            quote_codes.append(code)
    # quote_codes = [code for code in catalog.quotes.keys()
    #                if catalog.quotes[code].chapter == chapter and
    #                catalog.quotes[code].collection == collection and
    #                catalog.quotes[code].section == section and
    #                catalog.quotes[code].subsection == subsection and
    #                catalog.quotes[code].table == table
    #                ]
    if len(quote_codes) > 0:
        return sorted(quote_codes, key=lambda x: tuple(map(int, get_numeric_stamp(x, 'quote'))))
    return None


def quotes_output(sheet: worksheet, catalog: Catalog, start_line: int,
                  chapter: str = None, collection: str = None,
                  section: str = None, subsection: str = None, table: str = None
                  ) -> int:
    """ Записывает информацию о Расценках из catalog на лист sheet на строку row. """
    quotes = _get_quotes(catalog, chapter, collection, section, subsection, table)
    if quotes:
        row = start_line
        for quote in quotes:
            # print('\t\t\t\t\t', f"{catalog.quotes[quote].code!r} {catalog.quotes[quote].title}")
            row = _quote_line_output(quote, catalog, sheet, row)

        return row
    # print(f"\t\t\t\t\tнет ни одной расценки")
    return start_line
