from openpyxl.worksheet import worksheet
from openpyxl.styles import Font, Color
from openpyxl.utils.cell import column_index_from_string
from openpyxl.styles import numbers


from settings import Catalog, items_fonts, item_index
from catalog import get_numeric_stamp
from excelutils.quotes_output import quotes_output
from excelutils.tables_output import tables_output
from excelutils.sections_output import sections_output
from excelutils.subsections_output import subsections_output


def _collection_line_output(code: str, catalog: Catalog, sheet: worksheet, row: int) -> int:
    """ Записывает информацию из каталога о Сборнике с шифром code на лист sheet в строку row. """
    collection = catalog.collections[code]
    sheet.cell(row=row, column=column_index_from_string('A')).value = collection.chapter
    sheet.cell(row=row, column=column_index_from_string('B')).value = collection.code
    sheet.cell(row=row, column=column_index_from_string('G')).value = collection.title
    # группировка 2
    group_number = item_index['collection']+1
    sheet.row_dimensions.group(row, row+1, outline_level=group_number)

    columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G']
    for column in columns:
        sheet.cell(row=row, column=column_index_from_string(column)).style = 'chapter_line'
        sheet.cell(row=row, column=column_index_from_string(column)).font = items_fonts[item_index['collection']]
        sheet.cell(row=row, column=column_index_from_string(column)).number_format = numbers.FORMAT_TEXT
    return row+1


def _get_collection(catalog: Catalog = None, chapter: str = None) -> list[str] | None:
    """ Формирует список шифров сборников для указанной главы. """
    collection_codes = [code for code in catalog.collections.keys()
                        if catalog.collections[code].chapter == chapter]
    if len(collection_codes) > 0:
        return sorted(collection_codes, key=lambda x: tuple(map(int, get_numeric_stamp(x, 'collection'))))
    return None


def collection_output(sheet: worksheet, catalog: Catalog, start_line: int, chapter: str = None) -> int:
    """ Записывает информацию из каталога о сборниках на лист sheet начиная со строки start_line. """
    collections = _get_collection(catalog, chapter)
    # row = start_line
    if collections:
        row = start_line
        for collection in collections:
            # print('\t', f"{catalog.collections[collection].code!r} {catalog.collections[collection].title}")
            row = _collection_line_output(collection, catalog, sheet, row=row)
            row = quotes_output(sheet, catalog, row, chapter=chapter, collection=collection)
            row = tables_output(sheet, catalog, row, chapter=chapter, collection=collection)
            row = subsections_output(sheet, catalog, row, chapter=chapter, collection=collection)
            row = sections_output(sheet, catalog, row, chapter=chapter, collection=collection)
        return row
    # print(f"\tнет ни одного сборника")
    return start_line
