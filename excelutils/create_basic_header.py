from openpyxl.worksheet import worksheet
from openpyxl.utils.cell import column_index_from_string

from settings.excel_layout import headers, width_columns


def create_basic_header(sheet: worksheet):
    sheet.append(["."])
    headers["A:O"].extend(headers["P:T"])
    sheet.append(headers["A:O"])
    for column in range(1, len(headers["A:O"]) + 1):
        sheet.cell(row=2, column=column).style = 'title_basic'
    sheet.cell(row=1, column=column_index_from_string('K')).value = headers['K1']
    sheet.cell(row=1, column=column_index_from_string('K')).style = 'title_basic'
    sheet.merge_cells('K1:M1')

    sheet.cell(row=1, column=column_index_from_string('N')).value = headers['N1']
    sheet.cell(row=1, column=column_index_from_string('N')).style = 'title_basic'
    sheet.merge_cells('N1:O1')

    sheet.cell(row=1, column=column_index_from_string('P')).value = headers['P1']
    sheet.cell(row=1, column=column_index_from_string('P')).style = 'title_basic'
    sheet.merge_cells('P1:T1')

    for width in width_columns:
        sheet.column_dimensions[width].width = width_columns[width]
