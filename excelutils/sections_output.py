from openpyxl.worksheet import worksheet
from openpyxl.utils.cell import column_index_from_string
from openpyxl.styles import numbers

from settings import Catalog, items_fonts, item_index
from catalog import get_numeric_stamp
from excelutils.quotes_output import quotes_output
from excelutils.tables_output import tables_output
from excelutils.subsections_output import subsections_output


def _section_line_output(code: str, catalog: Catalog, sheet: worksheet, row: int) -> int:
    """ Записывает информацию из каталога об Отделе с шифром code на лист sheet в строку row. """
    section = catalog.sections[code]
    sheet.cell(row=row, column=column_index_from_string('A')).value = section.chapter
    sheet.cell(row=row, column=column_index_from_string('B')).value = section.collection
    sheet.cell(row=row, column=column_index_from_string('C')).value = section.code
    sheet.cell(row=row, column=column_index_from_string('G')).value = section.title
    # группировка 3
    group_number = item_index['section']+1
    sheet.row_dimensions.group(row, row+1, outline_level=group_number)

    columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G']
    for column in columns:
        sheet.cell(row=row, column=column_index_from_string(column)).style = 'chapter_line'
        sheet.cell(row=row, column=column_index_from_string(column)).font = items_fonts[item_index['section']]
        sheet.cell(row=row, column=column_index_from_string(column)).number_format = numbers.FORMAT_TEXT
    return row+1


def _get_section(catalog: Catalog = None,
                 chapter: str = None, collection: str = None) -> list[str] | None:
    """ Формирует список шифров Отделов для указанной главы и сборника. """
    section_codes = [
        code for code in catalog.sections.keys()
        if catalog.sections[code].chapter == chapter and catalog.sections[code].collection == collection
    ]
    if len(section_codes) > 0:
        return sorted(section_codes, key=lambda x: tuple(map(int, get_numeric_stamp(x, 'section'))))
    return None


def sections_output(sheet: worksheet, catalog: Catalog, start_line: int,
                    chapter: str = None, collection: str = None) -> int:
    """ Записывает информацию из каталога об Отделах на лист sheet начиная со строки start_line. """
    sections = _get_section(catalog, chapter, collection)
    if sections:
        row = start_line
        for section in sections:
            # print('\t\t', f"{catalog.sections[section].code!r} {catalog.sections[section].title}")
            row = _section_line_output(section, catalog, sheet, row)
            row = quotes_output(sheet, catalog, row, chapter=chapter, collection=collection, section=section)
            row = tables_output(sheet, catalog, row, chapter=chapter, collection=collection, section=section)
            row = subsections_output(sheet, catalog, row, chapter=chapter, collection=collection, section=section)
        return row
    # print(f"\t\tнет ни одного отдела")
    return start_line
