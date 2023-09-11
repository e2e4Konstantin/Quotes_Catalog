from openpyxl.worksheet import worksheet
from openpyxl.utils.cell import column_index_from_string
from openpyxl.styles import numbers

from settings import Catalog, items_fonts, item_index
from catalog import get_numeric_stamp

from excelutils.quotes_output import quotes_output
from excelutils.tables_output import tables_output


def _subsection_line_output(code: str, catalog: Catalog, sheet: worksheet, row: int) -> int:
    """ Записывает информацию из каталога об Отделе с шифром code на лист sheet в строку row. """
    subsection = catalog.subsections[code]
    sheet.cell(row=row, column=column_index_from_string('A')).value = subsection.chapter
    sheet.cell(row=row, column=column_index_from_string('B')).value = subsection.collection
    sheet.cell(row=row, column=column_index_from_string('C')).value = subsection.section
    sheet.cell(row=row, column=column_index_from_string('D')).value = subsection.code
    sheet.cell(row=row, column=column_index_from_string('G')).value = subsection.title
    # группировка 4
    group_number = item_index['subsection'] + 1
    sheet.row_dimensions.group(row, row + 1, outline_level=group_number)

    columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G']
    for column in columns:
        sheet.cell(row=row, column=column_index_from_string(column)).style = 'chapter_line'
        sheet.cell(row=row, column=column_index_from_string(column)).font = items_fonts[item_index['subsection']]
        sheet.cell(row=row, column=column_index_from_string(column)).number_format = numbers.FORMAT_TEXT
    return row + 1


def _get_subsection(catalog: Catalog = None,
                    chapter: str = None, collection: str = None, section: str = None) -> list[str] | None:
    """ Формирует список шифров Разделов для указанной главы, сборника и отдела. """
    subsection_codes = [
        code for code in catalog.subsections.keys()
        if catalog.subsections[code].chapter == chapter and
           catalog.subsections[code].collection == collection and
           catalog.subsections[code].section == section
    ]
    if len(subsection_codes) > 0:
        return sorted(subsection_codes, key=lambda x: tuple(map(int, get_numeric_stamp(x, 'subsection'))))
    return None


def subsections_output(sheet: worksheet, catalog: Catalog, start_line: int,
                       chapter: str = None, collection: str = None, section: str = None) -> int:
    """ Записывает информацию из каталога об Разделах на лист sheet начиная со строки start_line. """
    subsections = _get_subsection(catalog, chapter, collection, section)
    if subsections:
        row = start_line
        for subsection in subsections:
            # print('\t\t\t', f"{catalog.subsections[subsection].code!r} {catalog.subsections[subsection].title}")
            row = _subsection_line_output(subsection, catalog, sheet, row)
            row = quotes_output(sheet, catalog, row, chapter, collection, section, subsection)
            row = tables_output(sheet, catalog, row, chapter, collection, section, subsection)
        return row
    # print(f"\t\t\tнет ни одного раздела")
    return start_line
