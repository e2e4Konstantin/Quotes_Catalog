from openpyxl.worksheet import worksheet
from openpyxl.styles import (NamedStyle, Font, Border, Side, PatternFill, Alignment, Color)
from openpyxl.utils.cell import column_index_from_string
from settings import Catalog, items_fonts, item_index
from catalog import get_numeric_stamp

from excelutils.quotes_output import quotes_output
from excelutils.tables_output import tables_output
from excelutils.sections_output import sections_output
from excelutils.subsections_output import subsections_output
from excelutils.collections_output import collection_output


def _get_chapters(catalog: Catalog = None) -> list[str] | None:
    """ Готовит список глав из каталога и сортирует его по номеру главы """
    chapter_codes = list(catalog.chapters.keys())
    if len(chapter_codes) > 0:
        chapter_codes.sort(key=lambda code: int(get_numeric_stamp(code, 'chapter')[0]))
        return chapter_codes
    return None


def _range_decorating(sheet: worksheet, row, columns: list[str], style_name: str):
    """ Устанавливает стиль style_name для ячеек из списка columns """
    for column in columns:
        sheet.cell(row=row, column=column_index_from_string(column)).style = style_name
        sheet.cell(row=row, column=column_index_from_string(column)).font = items_fonts[item_index['chapter']]


def _chapter_line_output(code: str, catalog: Catalog, sheet: worksheet, row: int) -> int:
    """ Записывает информацию из каталога о главе code на лист sheet в строку row. """
    chapter = catalog.chapters[code]
    sheet.cell(row=row, column=column_index_from_string('A')).value = chapter.code
    sheet.cell(row=row, column=column_index_from_string('G')).value = chapter.title
    # группировка 1
    group_number = item_index['chapter']+1
    sheet.row_dimensions.group(row, row+2, outline_level=group_number)

    _range_decorating(sheet, row, ['A', 'B', 'C', 'D', 'E', 'F', 'G'], 'chapter_line')
    return row+1


def chapters_output(sheet: worksheet, catalog: Catalog, start_line: int, limit: int = 0) -> bool:
    """ Записывает информацию из каталога о главах на лист sheet начиная со строки row. """

    chapters = _get_chapters(catalog)
    if chapters:
        if 0 < limit <= len(chapters):
            chapters = chapters[: limit]
        # print(f"главы ({len(chapters)}): {chapters}")
        row = start_line
        for chapter in chapters:
            row = _chapter_line_output(chapter, catalog, sheet, row)
            # print(f"{catalog.chapters[chapter].code!r} {catalog.chapters[chapter].title}")
            row = quotes_output(sheet, catalog, row, chapter=chapter)
            row = tables_output(sheet, catalog, row, chapter=chapter)
            row = subsections_output(sheet, catalog, row, chapter=chapter)
            row = sections_output(sheet, catalog, row, chapter=chapter)
            row = collection_output(sheet, catalog, row, chapter=chapter)
            # row += 1
        return True
    # print(f"нет ни одной Главы")
    return False
