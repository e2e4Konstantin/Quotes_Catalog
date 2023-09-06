import sys
import openpyxl
from openpyxl.styles import (NamedStyle, Font, Border, Side, PatternFill, Alignment)

from filesutils import does_file_in_use
from settings import console_colors
from settings.excel_layout import cell_styles



class ExcelFile:
    def __init__(self, full_file_name: str = None):
        self.file_name = full_file_name
        self.sheet_name = None
        self.book = None
        self.sheet = None
        self.open_file()

    def __enter__(self):
        """ Вызывается при старте контекстного менеджера. """
        return self if self.book else None

    def __exit__(self, exception_type, exception_value, traceback):
        """ Будет вызван в завершении конструкции with, или в случае возникновения ошибки после нее. """
        self.close_file()

    def __str__(self):
        return f"excel file: '{self.file_name}', sheet: '{self.sheet_name}', link_sheet: '{self.sheet}'"

    def open_file(self):
        """ Создает новый excel файл. """
        try:
            self.book = openpyxl.Workbook()     # создаем экземпляр класса
        except IOError as err:
            error_out = f"{console_colors['RED']}{err}{console_colors['RESET']}"
            print(f"Ошибка при создании файла: '{self.file_name}'\n\t{error_out}")
            sys.exit()

    def close_file(self):
        if does_file_in_use(self.file_name):
            print(f"Не могу записать файл {self.file_name}, он используется другим приложением.")
        else:
            if self.book:
                self.book.save(self.file_name)
                self.book.close()
                self.sheet_name = None
                self.book = None
                self.sheet = None

    def create_sheets(self, sheets_name: list[str] = None):
        if self.book and sheets_name:
            for name in sheets_name:
                self.book.create_sheet(name)
            for sheet in self.book.worksheets:
                if sheet.title not in sheets_name:
                    self.book.remove(sheet)

    def set_sheet_grid(self, grid: bool = True):
        if self.book and self.sheet:
            self.sheet.sheet_view.showGridLines = grid

    def styles_init(self):
        name_styles = {}
        for item in cell_styles:
            name_styles[item['name']] = NamedStyle(name=item['name'])
            # header_style = NamedStyle(name=item['name'])
            name_styles[item['name']].font = Font(name=item['font']['name'], bold=item['font']['bold'],
                                                  size=item['font']['size'])
            bd = Side(style=item['border']['style'], color=item['border']['color'])
            name_styles[item['name']].border = Border(left=bd, top=bd, right=bd, bottom=bd)
            name_styles[item['name']].fill = PatternFill(patternType=item['fill']['patternType'],
                                                         fgColor=item['fill']['fgColor'])
            name_styles[item['name']].alignment = Alignment(horizontal=item['alignment']['horizontal'],
                                                            vertical=item['alignment']['vertical'],
                                                            wrap_text=item['alignment']['wrap_text'],
                                                            shrink_to_fit=item['alignment']['shrink_to_fit'],
                                                            indent=item['alignment']['indent'])
            if not (item['name'] in self.book.named_styles):
                self.book.add_named_style(name_styles[item['name']])



if __name__ == "__main__":
    full_name = r"F:\Kazak\GoogleDrive\1_KK\Job_CNAC\Python_projects\Quote_Catalog\output\output_template.xlsx"
    sheets = ["name", "stat"]
    file = ExcelFile(full_name)
    with file as ex:
        ex.create_sheets(sheets)
    # ex.close_file()
